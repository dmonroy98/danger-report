import os
import pandas as pd
from flask import Flask, render_template, jsonify, request, redirect, flash, make_response
from openpyxl import load_workbook

app = Flask(__name__)

# Secret key for flash messages
app.secret_key = os.getenv("FLASK_SECRET_KEY", "dev-secret")

# Password stored securely (e.g. in Render environment variables)
UPLOAD_PASSWORD = os.getenv("UPLOAD_PASSWORD")

# Local file path
LOCAL_FILE_PATH = "data/Danger Report Master.xlsm"

# Day mapping for sorting
DAY_ORDER = {
    "M": 1,
    "Tu": 2,
    "W": 3,
    "Th": 4,
    "F": 5,
    "Sa": 6,
    "Su": 7
}

# Row color classes
DAY_COLOR = {
    "M": "day-m",
    "Tu": "day-tu",
    "W": "day-w",
    "Th": "day-th",
    "F": "day-f",
    "Sa": "day-sa",
    "Su": "day-su"
}


def extract_day_code(class_name):
    if not isinstance(class_name, str):
        return ""
    parts = class_name.strip().split()
    if parts:
        return parts[-1]
    return ""


def load_excel_from_local():
    if not os.path.exists(LOCAL_FILE_PATH):
        raise FileNotFoundError(f"Excel file not found at: {LOCAL_FILE_PATH}")
    wb = load_workbook(filename=LOCAL_FILE_PATH, data_only=True)
    return wb


@app.route("/")
def home():
    return render_template("main.html")


@app.route("/danger-report", methods=["GET", "POST"])
def danger_report():
    try:
        wb = load_excel_from_local()
    except Exception as e:
        return f"Error loading Excel file: {str(e)}. Please upload a valid file."

    # Clean sheet names
    sheet_names = [str(s).strip() for s in wb.sheetnames]
    instructors = [s for s in sheet_names if "combined" not in s.lower()]

    if not instructors:
        return render_template(
            "danger_report.html",
            instructors=[],
            current=None,
            table="<p style='color:red'>No instructor sheets found in the workbook.</p>"
        )

    # Default to first instructor
    current = instructors[0]

    # Get selected instructor from URL (?instructor=...) or form
    selected = request.args.get("instructor") or request.form.get("instructor")
    if selected and selected in instructors:
        current = selected

    table_html = "<p>No data loaded.</p>"

    if current:
        try:
            ws = wb[current]
            data = list(ws.values)

            print(f"\nDEBUG: Sheet '{current}' - total rows read: {len(data)}")
            print("DEBUG: First 15 rows preview (first 6 columns):")
            for i, row in enumerate(data[:15], 1):
                preview = [str(cell)[:50].replace("\n", " ") if cell is not None else "" for cell in row[:6]]
                print(f"  Row {i:2d}: {preview}")

            # --- Header detection ---
            header_index = None
            for i, row in enumerate(data[:25]):
                row_clean = [str(cell).strip() for cell in row if cell is not None]
                row_lower = [s.lower() for s in row_clean]

                # Strict match
                if "Class Name" in row_clean and "Instructors" in row_clean:
                    header_index = i
                    print(f"DEBUG: Exact header match found on row {i+1}")
                    break

                # Loose/partial match fallback
                has_class = any("class" in s and "name" in s for s in row_lower)
                has_instructor = any("instruct" in s for s in row_lower)
                if has_class and has_instructor:
                    header_index = i
                    print(f"DEBUG: Partial header match found on row {i+1}")
                    break

            if header_index is not None:
                # Build DataFrame from header row onward
                df = pd.DataFrame(data[header_index:], columns=data[header_index])
                df.columns = [str(c).strip() if c is not None else "" for c in df.columns]
                df = df.iloc[1:].reset_index(drop=True)  # drop header row from data

                # Clean up
                if "Instructors" in df.columns:
                    df["Instructors"] = df["Instructors"].astype(str).str.strip()

                if "Class Name" in df.columns:
                    df["__day_code"] = df["Class Name"].apply(extract_day_code)
                    df["__day_sort"] = df["__day_code"].map(DAY_ORDER).fillna(999)
                    df["__day_color"] = df["__day_code"].map(DAY_COLOR).fillna("")

                table_html = df.to_html(classes="danger-table", index=False, na_rep="", escape=False)
            else:
                table_html = "<p style='color:red'>Could not find table headers ('Class Name' + 'Instructors') in first 25 rows.</p>"
                # Show preview of beginning of sheet
                preview_rows = []
                for row in data[:8]:
                    cells = [str(c)[:60].replace("\n", " ") if c is not None else "" for c in row[:6]]
                    preview_rows.append(", ".join(cells))
                table_html += "<pre style='font-size:0.9em; background:#f8f8f8; padding:10px;'>" + \
                              "\n".join(preview_rows) + "</pre>"

        except Exception as e:
            print(f"ERROR processing sheet '{current}': {str(e)}")
            table_html = f"<p style='color:red'>Error reading sheet '{current}': {str(e)}</p>"

    response = make_response(render_template(
        "danger_report.html",
        instructors=instructors,
        current=current,
        table=table_html
    ))
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response


# API Endpoints
@app.route("/api/get-sheets")
def get_sheets():
    try:
        wb = load_excel_from_local()
        sheet_names = [s.strip() for s in wb.sheetnames]
        filtered = [s for s in sheet_names if "combined" not in s.lower()]
        return jsonify({"sheets": filtered})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/get-sheet-data/<sheet_name>")
def get_sheet_data(sheet_name):
    try:
        wb = load_excel_from_local()
        if sheet_name not in wb.sheetnames:
            return jsonify({"error": f"Sheet '{sheet_name}' not found"}), 404

        ws = wb[sheet_name]
        data = list(ws.values)

        # Very basic version — you may want to reuse the improved header logic here later
        if len(data) < 2:
            return jsonify({"columns": [], "rows": []})

        df = pd.DataFrame(data[1:], columns=[str(c).strip() if c else "" for c in data[0]])

        if "Instructors" in df.columns:
            df["Instructors"] = df["Instructors"].astype(str).str.strip()

        if "Class Name" in df.columns:
            df["__day_code"] = df["Class Name"].apply(extract_day_code)
            df["__day_sort"] = df["__day_code"].map(DAY_ORDER).fillna(999)
            df["__day_color"] = df["__day_code"].map(DAY_COLOR).fillna("")

        result = {
            "columns": list(df.columns),
            "rows": df.fillna("").values.tolist()
        }
        return jsonify(result)

    except Exception as e:
        print(f"API error for sheet {sheet_name}: {str(e)}")
        return jsonify({"error": str(e)}), 500


# Upload routes
@app.route("/upload")
def upload_page():
    return render_template("upload.html")


@app.route("/upload-file", methods=["POST"])
def upload_file():
    password = request.form.get("password")
    file = request.files.get("file")

    if password != UPLOAD_PASSWORD:
        flash("Invalid password", "error")
        return redirect("/upload")

    if not file:
        flash("No file uploaded", "error")
        return redirect("/upload")

    filename = file.filename.lower()
    if not filename.endswith((".xlsm", ".xlsx")):
        flash("Invalid file type. Must be .xlsm or .xlsx", "error")
        return redirect("/upload")

    data_dir = os.path.dirname(LOCAL_FILE_PATH)
    os.makedirs(data_dir, exist_ok=True)

    # Optional backup
    if os.path.exists(LOCAL_FILE_PATH):
        try:
            os.replace(LOCAL_FILE_PATH, LOCAL_FILE_PATH + ".backup")
        except Exception as e:
            print(f"Backup failed: {e}")

    try:
        file.save(LOCAL_FILE_PATH)
        flash("File uploaded successfully!", "success")
    except Exception as e:
        flash(f"Failed to save file: {str(e)}", "error")
        return redirect("/upload")

    return redirect("/danger-report")


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)